import os
import re
import json
import time
from typing import Dict, List, Tuple

import openpyxl
from openai import OpenAI

INPUT_PATH = os.environ.get(
    "INPUT_XLSX",
    "JLPT_N3_N5_WORD_KANJI_A_영어포함_한국어빈칸_품사완성.xlsx"
)
OUTPUT_PATH = os.environ.get(
    "OUTPUT_XLSX",
    "JLPT_N3_N5_WORD_KANJI_A_한국어완성.xlsx"
)

MODEL = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")  # Structured outputs compatible per docs.

BATCH_SIZE = int(os.environ.get("BATCH_SIZE", "50"))
SLEEP_SEC = float(os.environ.get("SLEEP_SEC", "0.2"))

ENGLISH_RE = re.compile(r"[A-Za-z]")

def is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False

def has_english(text: str) -> bool:
    return bool(ENGLISH_RE.search(text or ""))

def build_schema() -> dict:
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "translations": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "i": {"type": "integer"},
                        "ko": {"type": "string"},
                    },
                    "required": ["i", "ko"],
                },
            }
        },
        "required": ["translations"],
    }

def translate_batch(client: OpenAI, items: List[Tuple[int, str]]) -> Dict[int, str]:
    """
    items: [(row_index, english_text)]
    returns: {row_index: korean_text}
    """
    # Keep it simple and deterministic.
    payload = [{"i": i, "en": t} for i, t in items]
    schema = build_schema()

    for attempt in range(5):
        try:
            resp = client.responses.create(
                model=MODEL,
                instructions=(
                    "You are a professional Korean translator.\n"
                    "Translate the 'en' field into natural Korean.\n"
                    "Rules:\n"
                    "- Keep meaning accurate and concise (dictionary gloss style).\n"
                    "- If 'en' contains multiple senses separated by ';' or ',', keep separators and translate each sense.\n"
                    "- Do not add explanations.\n"
                    "- Return ONLY the JSON that matches the schema."
                ),
                input=f"Translate these items to Korean:\n{json.dumps(payload, ensure_ascii=False)}",
                text={
                    "format": {
                        "type": "json_schema",
                        "name": "translations_schema",
                        "strict": True,
                        "schema": schema,
                    }
                },
            )
            data = json.loads(resp.output_text)
            out: Dict[int, str] = {}
            for obj in data.get("translations", []):
                out[int(obj["i"])] = (obj["ko"] or "").strip()
            # basic sanity
            if len(out) != len(items):
                raise ValueError(f"Expected {len(items)} items, got {len(out)}")
            return out
        except Exception as e:
            if attempt == 4:
                raise
            time.sleep(1.0 * (attempt + 1))
    raise RuntimeError("unreachable")

def process_sheet(ws, client: OpenAI, sheet_name: str):
    # Find columns
    header = [c.value for c in ws[1]]
    try:
        col_en = header.index("영어뜻") + 1
        col_ko = header.index("한국어뜻") + 1
    except ValueError:
        raise RuntimeError(f"[{sheet_name}] '영어뜻'/'한국어뜻' 컬럼을 찾지 못했습니다: {header}")

    # Collect rows that need work
    to_translate: List[Tuple[int, str]] = []
    to_copy: List[int] = []

    for r in range(2, ws.max_row + 1):
        en = ws.cell(r, col_en).value
        ko = ws.cell(r, col_ko).value

        if is_blank(ko):
            if is_blank(en):
                # nothing to translate -> leave blank
                continue
            en_text = str(en).strip()
            if has_english(en_text):
                to_translate.append((r, en_text))
            else:
                # already Korean (or non-latin) -> copy
                to_copy.append(r)

    # Copy non-english directly
    for r in to_copy:
        ws.cell(r, col_ko).value = str(ws.cell(r, col_en).value).strip()

    # Translate in batches
    total = len(to_translate)
    if total == 0:
        return

    print(f"[{sheet_name}] translating {total} rows...")

    for start in range(0, total, BATCH_SIZE):
        batch = to_translate[start:start + BATCH_SIZE]
        mapping = translate_batch(client, batch)
        for r, _ in batch:
            ws.cell(r, col_ko).value = mapping.get(r, "").strip()
        done = min(start + BATCH_SIZE, total)
        print(f"[{sheet_name}] {done}/{total}")
        time.sleep(SLEEP_SEC)

def main():
    if not os.path.exists(INPUT_PATH):
        raise FileNotFoundError(f"INPUT_XLSX not found: {INPUT_PATH}")

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY 환경변수가 필요합니다.")

    client = OpenAI(api_key=api_key)

    wb = openpyxl.load_workbook(INPUT_PATH)
    for name in ("WORD", "KANJI"):
        if name not in wb.sheetnames:
            raise RuntimeError(f"시트를 찾지 못했습니다: {name}")
        process_sheet(wb[name], client, name)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
