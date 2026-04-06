import json
import os
from typing import Dict, List

import openpyxl

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DATA_DIR = os.path.join(BASE_DIR, "data")

INPUT_XLSX = os.environ.get("INPUT_XLSX")
OUTPUT_JSON = os.environ.get("OUTPUT_JSON", os.path.join(DATA_DIR, "jlpt_quiz_data.json"))
PLACEHOLDER = "아직없음"


def _pick_input_xlsx() -> str:
    if INPUT_XLSX:
        if os.path.isabs(INPUT_XLSX):
            return INPUT_XLSX
        return os.path.join(DATA_DIR, INPUT_XLSX)

    candidates = sorted(
        f for f in os.listdir(DATA_DIR) if f.lower().endswith(".xlsx")
    )
    if not candidates:
        raise FileNotFoundError(f"No .xlsx file found in {DATA_DIR}")
    return os.path.join(DATA_DIR, candidates[0])


def _safe_text(value) -> str:
    if value is None:
        return PLACEHOLDER
    text = str(value).strip()
    return text if text else PLACEHOLDER


def _bucket_key(level: str, kind: str) -> str:
    return f"{level.lower()}_{kind}"


def _build_from_workbook(xlsx_path: str) -> Dict[str, List[dict]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    data: Dict[str, List[dict]] = {
        "n5_kanji": [],
        "n4_kanji": [],
        "n3_kanji": [],
        "n5_word": [],
        "n4_word": [],
        "n3_word": [],
    }

    if "WORD" not in wb.sheetnames or "KANJI" not in wb.sheetnames:
        raise RuntimeError("Workbook must contain WORD and KANJI sheets.")

    word_ws = wb["WORD"]
    for row in word_ws.iter_rows(min_row=2, values_only=True):
        level = str(row[0]).strip().upper() if row[0] is not None else ""
        if level not in {"N5", "N4", "N3"}:
            continue
        item = {
            "surface": _safe_text(row[1]),
            "reading": _safe_text(row[2]),
            "meaning": _safe_text(row[5]),
            "pos": _safe_text(row[4]),
            "level": level,
            "kind": "word",
        }
        data[_bucket_key(level, "word")].append(item)

    kanji_ws = wb["KANJI"]
    for row in kanji_ws.iter_rows(min_row=2, values_only=True):
        level = str(row[0]).strip().upper() if row[0] is not None else ""
        if level not in {"N5", "N4", "N3"}:
            continue
        item = {
            "surface": _safe_text(row[1]),
            "reading": _safe_text(row[2]),
            "meaning": _safe_text(row[4]),
            "pos": "한자",
            "level": level,
            "kind": "kanji",
        }
        data[_bucket_key(level, "kanji")].append(item)

    return data


def main() -> None:
    input_path = _pick_input_xlsx()
    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)

    data = _build_from_workbook(input_path)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Input: {input_path}")
    print(f"Saved: {OUTPUT_JSON}")
    for key, items in data.items():
        print(f"{key}: {len(items)}")


if __name__ == "__main__":
    main()
